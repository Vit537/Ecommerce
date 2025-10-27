import csv
import io
from typing import Any, Dict, Iterable, List, Tuple
from datetime import datetime

from django.contrib.contenttypes.models import ContentType
from django.db.models import Model, QuerySet, Q
from django.utils import timezone
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.http import HttpResponse

from tenant.middleware import get_current_tenant


def _tenant_filtered_queryset(model: type[Model]) -> QuerySet:
    manager = model._default_manager
    qs = manager.all()
    tenant = get_current_tenant()
    if tenant and hasattr(model, 'tenant_id'):
        qs = qs.filter(tenant=tenant)
    return qs


def build_queryset(template_filters: Dict[str, Any], runtime_filters: Dict[str, Any], model: type[Model]) -> QuerySet:
    queryset = _tenant_filtered_queryset(model)
    filters = {**(template_filters or {}), **(runtime_filters or {})}
    if filters:
        queryset = queryset.filter(**filters)
    return queryset


def apply_ordering(queryset: QuerySet, ordering: Iterable[str] | None) -> QuerySet:
    if ordering:
        return queryset.order_by(*ordering)
    return queryset


def extract_rows(queryset: QuerySet, fields: List[str], *, limit: int | None = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    column_fields = fields or []
    if not column_fields:
        column_fields = ['id']
    if limit:
        queryset = queryset[:limit]
    rows = list(queryset.values(*column_fields))
    return column_fields, rows


def export_to_csv(columns: List[str], rows: List[Dict[str, Any]]) -> Tuple[str, int]:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction='ignore')
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    content = buffer.getvalue().encode('utf-8')
    filename = f"reports/export_{timezone.now().strftime('%Y%m%d%H%M%S')}.csv"
    path = default_storage.save(filename, ContentFile(content))
    return path, len(content)


def build_dynamic_filter(filter_config: Dict[str, Any]) -> Q:
    """
    Construye un objeto Q de Django a partir de una configuración de filtro
    
    filter_config format:
    {
        "field": "status",
        "operator": "equals|contains|gt|gte|lt|lte|in|isnull|startswith|endswith",
        "value": "active"
    }
    """
    field = filter_config.get('field')
    operator = filter_config.get('operator', 'equals')
    value = filter_config.get('value')
    
    if not field:
        return Q()
    
    operator_mapping = {
        'equals': '',
        'contains': '__icontains',
        'startswith': '__istartswith',
        'endswith': '__iendswith',
        'gt': '__gt',
        'gte': '__gte',
        'lt': '__lt',
        'lte': '__lte',
        'in': '__in',
        'isnull': '__isnull',
        'exact': '__iexact',
    }
    
    lookup = f"{field}{operator_mapping.get(operator, '')}"
    
    if operator == 'isnull':
        return Q(**{lookup: value in ['true', True, 1, '1']})
    
    return Q(**{lookup: value})


def apply_dynamic_filters(queryset: QuerySet, filters: List[Dict[str, Any]]) -> QuerySet:
    """Aplica una lista de filtros dinámicos a un queryset"""
    q_objects = Q()
    
    for filter_config in filters:
        q_objects &= build_dynamic_filter(filter_config)
    
    if q_objects:
        queryset = queryset.filter(q_objects)
    
    return queryset


def apply_dynamic_ordering(queryset: QuerySet, ordering: List[Dict[str, Any]]) -> QuerySet:
    """
    Aplica ordenamiento dinámico
    
    ordering format:
    [{"field": "created_at", "direction": "desc|asc"}]
    """
    order_fields = []
    
    for order_config in ordering:
        field = order_config.get('field')
        direction = order_config.get('direction', 'asc')
        
        if field:
            prefix = '-' if direction == 'desc' else ''
            order_fields.append(f"{prefix}{field}")
    
    if order_fields:
        queryset = queryset.order_by(*order_fields)
    
    return queryset


def export_to_excel_response(columns: List[str], rows: List[Dict[str, Any]], filename: str = None) -> HttpResponse:
    """
    Exporta datos a Excel usando openpyxl
    Requiere: pip install openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column.replace('_', ' ').title()
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, column in enumerate(columns, start=1):
            value = row_data.get(column, '')
            # Formatear fechas
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Crear respuesta HTTP
    if not filename:
        filename = f"reporte_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_to_pdf_response(columns: List[str], rows: List[Dict[str, Any]], title: str = "Reporte", filename: str = None) -> HttpResponse:
    """
    Exporta datos a PDF usando reportlab
    Requiere: pip install reportlab
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        raise ImportError("reportlab no está instalado. Ejecuta: pip install reportlab")
    
    buffer = io.BytesIO()
    
    # Usar orientación horizontal si hay muchas columnas
    pagesize = landscape(letter) if len(columns) > 6 else letter
    doc = SimpleDocTemplate(buffer, pagesize=pagesize)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Preparar datos para la tabla
    table_data = []
    
    # Encabezados
    headers = [column.replace('_', ' ').title() for column in columns]
    table_data.append(headers)
    
    # Datos
    for row in rows[:1000]:  # Limitar a 1000 registros para PDF
        row_data = []
        for column in columns:
            value = row.get(column, '')
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M')
            row_data.append(str(value))
        table_data.append(row_data)
    
    # Crear tabla
    table = Table(table_data)
    
    # Estilos de tabla
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Crear respuesta HTTP
    if not filename:
        filename = f"reporte_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
