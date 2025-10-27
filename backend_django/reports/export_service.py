"""
Servicios para exportar reportes a PDF y Excel
"""

import os
from typing import List, Dict, Any
from datetime import datetime
import pandas as pd
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

import matplotlib
matplotlib.use('Agg')  # Backend sin interfaz gráfica
import matplotlib.pyplot as plt


class ReportExporter:
    """
    Clase para exportar reportes a diferentes formatos
    """
    
    @staticmethod
    def to_excel(data: List[Dict], filename: str, report_metadata: Dict = None) -> BytesIO:
        """
        Exporta datos a Excel con formato profesional
        
        Args:
            data: Lista de diccionarios con los datos
            filename: Nombre sugerido para el archivo
            report_metadata: Información adicional del reporte
            
        Returns:
            BytesIO con el archivo Excel
        """
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Crear archivo en memoria
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Escribir datos
            df.to_excel(writer, sheet_name='Datos', index=False, startrow=0)
            
            # Obtener workbook y worksheet para formato
            workbook = writer.book
            worksheet = writer.sheets['Datos']
            
            # Estilo del encabezado
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Estilo de celdas
            cell_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Aplicar estilo al encabezado
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = cell_border
            
            # Ajustar ancho de columnas automáticamente
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell.border = cell_border
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                # Ancho mínimo de 12, máximo de 60
                adjusted_width = min(max(max_length + 3, 12), 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Alternar colores de filas
            light_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                if row_idx % 2 == 0:
                    for cell in row:
                        if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                            cell.fill = light_fill
            
            # Si hay metadata, crear hoja adicional
            if report_metadata:
                metadata_df = pd.DataFrame([
                    ['Reporte', report_metadata.get('title', 'N/A')],
                    ['Generado por', report_metadata.get('user', 'N/A')],
                    ['Fecha', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                    ['Total registros', len(data)],
                    ['Consulta', report_metadata.get('prompt', 'N/A')],
                ])
                metadata_df.to_excel(writer, sheet_name='Información', index=False, header=False)
                
                # Formato de la hoja de metadata
                meta_worksheet = writer.sheets['Información']
                for row in meta_worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        if cell.column == 1:
                            cell.font = Font(bold=True, size=11)
                
                # Ajustar ancho de columnas de metadata
                meta_worksheet.column_dimensions['A'].width = 20
                meta_worksheet.column_dimensions['B'].width = 60
        
        output.seek(0)
        return output
    
    @staticmethod
    def to_pdf(
        data: List[Dict], 
        filename: str, 
        report_metadata: Dict = None,
        chart_data: Dict = None
    ) -> BytesIO:
        """
        Exporta datos a PDF con formato profesional
        
        Args:
            data: Lista de diccionarios con los datos
            filename: Nombre sugerido para el archivo
            report_metadata: Información adicional del reporte
            chart_data: Datos para generar gráficos
            
        Returns:
            BytesIO con el archivo PDF
        """
        
        output = BytesIO()
        
        # Determinar orientación según número de columnas
        df = pd.DataFrame(data)
        num_columns = len(df.columns)
        
        # Si hay más de 6 columnas, usar landscape
        pagesize = landscape(A4) if num_columns > 6 else A4
        
        # Crear documento PDF
        doc = SimpleDocTemplate(
            output,
            pagesize=pagesize,
            rightMargin=20,
            leftMargin=20,
            topMargin=40,
            bottomMargin=40
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1a365d'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#2d3748'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        
        # Contenido del PDF
        story = []
        
        # Título
        title = Paragraph(
            report_metadata.get('title', 'Reporte de Productos') if report_metadata else 'Reporte',
            title_style
        )
        story.append(title)
        story.append(Spacer(1, 0.15 * inch))
        
        # Metadata
        if report_metadata:
            metadata_text = f"""
            <b>Generado por:</b> {report_metadata.get('user', 'Sistema')}<br/>
            <b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>
            <b>Consulta:</b> {report_metadata.get('prompt', 'N/A')}<br/>
            <b>Total de registros:</b> {len(data)}
            """
            story.append(Paragraph(metadata_text, styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
        
        # Resumen (si existe)
        if report_metadata and 'summary' in report_metadata:
            story.append(Paragraph('Resumen Ejecutivo', heading_style))
            story.append(Paragraph(report_metadata['summary'], styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
        
        # Tabla de datos (SIN GRÁFICO - se quitó como solicitaste)
        if data:
            story.append(Paragraph('Datos Detallados', heading_style))
            story.append(Spacer(1, 0.1 * inch))
            
            # NO limitar columnas - mostrar TODAS
            # Limitar solo filas si son demasiadas (para PDF)
            max_rows = 100  # Aumentado a 100 registros
            show_all_message = False
            
            if len(df) > max_rows:
                df_display = df.head(max_rows)
                show_all_message = True
            else:
                df_display = df
            
            if show_all_message:
                story.append(Paragraph(
                    f'<i>Mostrando primeros {max_rows} registros de {len(data)} total. Descarga el Excel para ver todos.</i>',
                    styles['Italic']
                ))
                story.append(Spacer(1, 0.1 * inch))
            
            # Crear tabla con TODAS las columnas
            table_data = [df_display.columns.tolist()] + df_display.values.tolist()
            
            # Calcular ancho de columnas dinámicamente
            page_width = pagesize[0] - 40  # Ancho disponible
            col_widths = [page_width / len(df_display.columns)] * len(df_display.columns)
            
            # Ajustar tamaño de fuente según número de columnas
            if num_columns > 10:
                font_size = 6
                header_font_size = 7
            elif num_columns > 7:
                font_size = 7
                header_font_size = 8
            else:
                font_size = 8
                header_font_size = 9
            
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Estilo de tabla mejorado
            table.setStyle(TableStyle([
                # Encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a5568')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # Cuerpo
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), font_size),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ]))
            
            story.append(table)
        
        # Construir PDF
        doc.build(story)
        
        output.seek(0)
        return output
    
    @staticmethod
    def _generate_chart(chart_config: Dict, data: List[Dict]) -> BytesIO:
        """
        Genera un gráfico con matplotlib
        
        Args:
            chart_config: Configuración del gráfico
                {
                    'type': 'bar'|'line'|'pie',
                    'x_field': str,
                    'y_field': str,
                    'title': str
                }
            data: Datos para graficar
            
        Returns:
            BytesIO con la imagen del gráfico
        """
        
        try:
            df = pd.DataFrame(data)
            
            chart_type = chart_config.get('type', 'bar')
            x_field = chart_config.get('x_field')
            y_field = chart_config.get('y_field')
            title = chart_config.get('title', 'Gráfico')
            
            # Limitar datos si son muchos
            if len(df) > 20:
                df = df.head(20)
            
            fig, ax = plt.subplots(figsize=(10, 6))
            
            if chart_type == 'bar':
                ax.bar(df[x_field].astype(str), df[y_field])
                ax.set_xlabel(x_field.replace('_', ' ').title())
                ax.set_ylabel(y_field.replace('_', ' ').title())
                plt.xticks(rotation=45, ha='right')
                
            elif chart_type == 'line':
                ax.plot(df[x_field], df[y_field], marker='o')
                ax.set_xlabel(x_field.replace('_', ' ').title())
                ax.set_ylabel(y_field.replace('_', ' ').title())
                plt.xticks(rotation=45, ha='right')
                
            elif chart_type == 'pie':
                ax.pie(df[y_field], labels=df[x_field].astype(str), autopct='%1.1f%%')
            
            ax.set_title(title, fontsize=14, fontweight='bold')
            plt.tight_layout()
            
            # Guardar en buffer
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            print(f"Error generando gráfico: {e}")
            return None
